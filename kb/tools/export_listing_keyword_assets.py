#!/usr/bin/env python3
"""Export structured assets from inbox listing workbook.

This script reads the local-only workbook copied into the controlled source
package and writes CSV/JSONL assets for RAG, BI, and Agent workflows.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = ROOT / "kb/_sources/inbox-independent-site/originals/listing优化模板.xlsx"
OUT_DIR = ROOT / "kb/_sources/inbox-independent-site/structured/listing-keyword-assets"
SOURCE_ID = "inbox:S05"

KEYWORD_SHEET = "5.关键词库"
CONTEXT_SHEETS = [
    "1.竞品库",
    "2.卖点库",
    "3.评价分析",
    "4.竞品QA分析",
    "3.竞品图库",
    "竞品5点库",
    "竞品标题库",
    "标题框架",
    "5.自己卖点框架",
    "6.作图框架",
    "7.作图需求",
]

KEYWORD_HEADER_MAP = {
    "关键词": "keyword",
    "关键词翻译": "keyword_translation",
    "趋势": "trend_ref",
    "旺季": "peak_season",
    "淡季": "off_season",
    "综合": "composite_score",
    "月搜索量": "monthly_search_volume",
    "月购买量": "monthly_purchase_volume",
    "购买率": "purchase_rate",
    "SPR": "spr",
    "PPC竞价": "ppc_bid",
    "AC推荐词": "ac_recommended_keyword",
    "流量占比": "traffic_share",
    "流量词类型": "traffic_word_type_raw",
    "首个ASIN自然排名": "first_asin_organic_rank",
    "首个ASIN广告排名": "first_asin_ad_rank",
    "预估周曝光量": "estimated_weekly_impressions",
    "相关产品": "related_products",
    "相关ASIN": "related_asin",
    "ABA周排名": "aba_weekly_rank",
    "标题密度": "title_density",
    "商品数": "product_count",
    "供需比": "supply_demand_ratio",
    "广告竞品数": "ad_competitor_count",
    "点击集中度": "click_concentration",
    "前三ASIN转化总占比": "top3_asin_conversion_share",
    "建议竞价范围": "suggested_bid_range",
    "#1 前三ASIN": "top1_asin",
    "#1 点击共享": "top1_click_share",
    "#1 转化共享": "top1_conversion_share",
    "#2 前三ASIN": "top2_asin",
    "#2 点击共享": "top2_click_share",
    "#2 转化共享": "top2_conversion_share",
    "#3 前三ASIN": "top3_asin",
    "#3 点击共享": "top3_click_share",
    "#3 转化共享": "top3_conversion_share",
    "前十ASIN": "top10_asin",
}

STANDARD_KEYWORD_FIELDS = [
    "keyword",
    "keyword_translation",
    "trend_ref",
    "peak_season",
    "off_season",
    "composite_score",
    "monthly_search_volume",
    "monthly_purchase_volume",
    "purchase_rate",
    "spr",
    "ppc_bid",
    "traffic_share",
    "traffic_word_type_raw",
    "first_asin_organic_rank",
    "first_asin_ad_rank",
    "estimated_weekly_impressions",
    "related_products",
    "related_asin",
    "aba_weekly_rank",
    "title_density",
    "product_count",
    "supply_demand_ratio",
    "ad_competitor_count",
    "click_concentration",
    "top3_asin_conversion_share",
    "suggested_bid_range",
    "top1_asin",
    "top1_click_share",
    "top1_conversion_share",
    "top2_asin",
    "top2_click_share",
    "top2_conversion_share",
    "top3_asin",
    "top3_click_share",
    "top3_conversion_share",
    "top10_asin",
]

NUMERIC_FIELDS = {
    "composite_score",
    "monthly_search_volume",
    "monthly_purchase_volume",
    "purchase_rate",
    "spr",
    "traffic_share",
    "first_asin_organic_rank",
    "first_asin_ad_rank",
    "estimated_weekly_impressions",
    "aba_weekly_rank",
    "title_density",
    "product_count",
    "supply_demand_ratio",
    "ad_competitor_count",
    "click_concentration",
    "top3_asin_conversion_share",
    "top1_click_share",
    "top1_conversion_share",
    "top2_click_share",
    "top2_conversion_share",
    "top3_click_share",
    "top3_conversion_share",
}

KNOWN_TRAFFIC_TYPES = ["自然搜索词", "HR推荐词", "品牌广告词", "视频广告词", "SP广告词"]
ASIN_RE = re.compile(r"\b(B0[A-Z0-9]{8}|[0-9]{10})\b", re.IGNORECASE)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return "\n".join(line.rstrip() for line in text.split("\n")).strip()


def to_number(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value) if isinstance(value, float) and value.is_integer() else value
    text = clean_text(value)
    if not text:
        return None
    text = text.replace("$", "").replace(",", "").replace("%", "")
    text = text.replace("，", "")
    try:
        n = float(text)
    except ValueError:
        return None
    return int(n) if n.is_integer() else n


def csv_value(value: Any) -> Any:
    text = clean_text(value)
    if text and text[0] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def safe_header(value: Any, index: int, seen: dict[str, int]) -> str:
    base = clean_text(value) or f"unnamed_{index:02d}"
    base = re.sub(r"\s+", "_", base)
    count = seen.get(base, 0) + 1
    seen[base] = count
    return base if count == 1 else f"{base}_{count}"


def cell(values: list[Any], index: int) -> Any:
    return values[index - 1] if 0 <= index - 1 < len(values) else None


def looks_like_asin(value: Any) -> bool:
    return bool(ASIN_RE.search(clean_text(value)))


def looks_like_traffic_type(value: Any) -> bool:
    text = clean_text(value)
    return any(token in text for token in KNOWN_TRAFFIC_TYPES)


def split_traffic_types(raw: str) -> list[str]:
    if not raw:
        return []
    return [token for token in KNOWN_TRAFFIC_TYPES if token in raw]


def used_rows(ws) -> list[tuple[int, list[Any]]]:
    rows: list[tuple[int, list[Any]]] = []
    max_col = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = list(row)
        last = 0
        for i, value in enumerate(vals, 1):
            if value is not None and clean_text(value) != "":
                last = i
        if last:
            rows.append((row_idx, vals[:last]))
            max_col = max(max_col, last)
    normalized = []
    for row_idx, vals in rows:
        normalized.append((row_idx, vals + [None] * (max_col - len(vals))))
    return normalized


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: csv_value(row.get(k, "")) for k in headers})


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def export_sheet_csv(ws, output_dir: Path) -> dict[str, Any]:
    rows = used_rows(ws)
    if not rows:
        return {
            "sheet_name": ws.title,
            "row_count": 0,
            "column_count": 0,
            "export_path": "",
        }

    header_row_idx, first_row = rows[0]
    seen: dict[str, int] = {}
    headers = [safe_header(v, i, seen) for i, v in enumerate(first_row, 1)]
    data_rows = []
    nonempty_data_rows = 0
    for source_row, vals in rows[1:]:
        record = {"source_row": source_row}
        for header, value in zip(headers, vals):
            record[header] = value
        if any(clean_text(v) for v in vals):
            nonempty_data_rows += 1
        data_rows.append(record)
    export_path = output_dir / "sheet_exports" / f"{ws.title}.csv"
    write_csv(export_path, ["source_row"] + headers, data_rows)
    return {
        "sheet_name": ws.title,
        "header_row": header_row_idx,
        "data_row_count": nonempty_data_rows,
        "column_count": len(headers),
        "export_path": str(export_path.relative_to(OUT_DIR)),
    }


def normalized_keyword_headers(raw_headers: list[Any]) -> list[str]:
    unnamed_counter = 0
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, header in enumerate(raw_headers, 1):
        text = clean_text(header)
        if text:
            mapped = KEYWORD_HEADER_MAP.get(text, text)
        else:
            unnamed_counter += 1
            mapped = f"unnamed_{index:02d}"
        count = seen.get(mapped, 0) + 1
        seen[mapped] = count
        headers.append(mapped if count == 1 else f"{mapped}_{count}")
    return headers


def derive_keyword_tier(row: dict[str, Any]) -> str:
    search = to_number(row.get("monthly_search_volume")) or 0
    purchase = to_number(row.get("monthly_purchase_volume")) or 0
    if search >= 100000 or purchase >= 5000:
        return "head"
    if search >= 10000 or purchase >= 500:
        return "mid"
    return "long_tail"


def raw_columns(raw_header_labels: list[str], values: list[Any]) -> dict[str, str]:
    columns = {}
    for index, value in enumerate(values, 1):
        header = raw_header_labels[index - 1] if index - 1 < len(raw_header_labels) else ""
        label = re.sub(r"\s+", "_", header) if header else "unnamed"
        columns[f"col_{index:02d}_{label}"] = clean_text(value)
    return columns


def classify_keyword_layout(values: list[Any]) -> str:
    if looks_like_asin(cell(values, 33)):
        return "rich_top_asin_col33"
    if looks_like_asin(cell(values, 25)) and looks_like_traffic_type(cell(values, 6)):
        return "compact_type_col6_top_asin_col25"
    if looks_like_asin(cell(values, 25)):
        return "compact_metric_col6_top_asin_col25"
    if looks_like_traffic_type(cell(values, 14)):
        return "rich_no_top_asin"
    if looks_like_traffic_type(cell(values, 6)):
        return "compact_type_col6_no_top_asin"
    return "unknown"


def maybe_number(value: Any) -> Any:
    n = to_number(value)
    return "" if n is None else n


def best_ppc_bid(values: list[Any], *indexes: int) -> str:
    for index in indexes:
        value = clean_text(cell(values, index))
        if value.startswith("$"):
            return value
    for index in indexes:
        value = clean_text(cell(values, index))
        if value:
            return value
    return ""


def build_keyword_record(
    ordinal: int,
    source_row: int,
    values: list[Any],
    raw_header_labels: list[str],
    source_sha: str,
) -> dict[str, Any]:
    layout = classify_keyword_layout(values)
    is_rich = layout.startswith("rich")
    is_compact = layout.startswith("compact")
    top_start = 33 if looks_like_asin(cell(values, 33)) else 25 if looks_like_asin(cell(values, 25)) else None

    record: dict[str, Any] = {
        "keyword_asset_id": f"kw_{ordinal:04d}",
        "source_id": SOURCE_ID,
        "source_file": "listing优化模板.xlsx",
        "source_sha256_12": source_sha[:12],
        "source_sheet": KEYWORD_SHEET,
        "source_row": source_row,
        "row_layout": layout,
    }
    for field in STANDARD_KEYWORD_FIELDS:
        record[field] = ""

    record.update(
        {
            "keyword": clean_text(cell(values, 1)),
            "keyword_translation": clean_text(cell(values, 2)),
            "trend_ref": clean_text(cell(values, 3)),
            "peak_season": clean_text(cell(values, 4)),
            "off_season": clean_text(cell(values, 5)) if is_rich else "",
        }
    )

    if is_rich:
        record.update(
            {
                "composite_score": maybe_number(cell(values, 6)),
                "monthly_search_volume": maybe_number(cell(values, 7)),
                "monthly_purchase_volume": maybe_number(cell(values, 8)),
                "purchase_rate": maybe_number(cell(values, 9)),
                "spr": maybe_number(cell(values, 10)),
                "ppc_bid": best_ppc_bid(values, 11),
                "traffic_share": maybe_number(cell(values, 13)),
                "traffic_word_type_raw": clean_text(cell(values, 14))
                if looks_like_traffic_type(cell(values, 14))
                else "",
                "first_asin_organic_rank": maybe_number(cell(values, 15)),
                "first_asin_ad_rank": maybe_number(cell(values, 16)),
                "estimated_weekly_impressions": maybe_number(cell(values, 17)),
                "related_products": clean_text(cell(values, 18)),
                "related_asin": clean_text(cell(values, 19)),
                "aba_weekly_rank": maybe_number(cell(values, 20)),
                "title_density": maybe_number(cell(values, 25)),
                "product_count": maybe_number(cell(values, 26)),
                "supply_demand_ratio": maybe_number(cell(values, 27)),
                "ad_competitor_count": maybe_number(cell(values, 28)),
                "click_concentration": maybe_number(cell(values, 29)),
                "top3_asin_conversion_share": maybe_number(cell(values, 30)),
                "suggested_bid_range": clean_text(cell(values, 32)),
            }
        )
    elif is_compact:
        record.update(
            {
                "composite_score": maybe_number(cell(values, 6))
                if not looks_like_traffic_type(cell(values, 6))
                else "",
                "monthly_search_volume": maybe_number(cell(values, 7)),
                "monthly_purchase_volume": maybe_number(cell(values, 8)),
                "purchase_rate": maybe_number(cell(values, 9)),
                "spr": maybe_number(cell(values, 10)),
                "ppc_bid": best_ppc_bid(values, 23, 11),
                "traffic_share": maybe_number(cell(values, 5)),
                "traffic_word_type_raw": clean_text(cell(values, 6))
                if looks_like_traffic_type(cell(values, 6))
                else "",
                "aba_weekly_rank": maybe_number(cell(values, 12)),
                "product_count": maybe_number(cell(values, 18)),
                "supply_demand_ratio": maybe_number(cell(values, 19)),
                "ad_competitor_count": maybe_number(cell(values, 20)),
                "click_concentration": maybe_number(cell(values, 21)),
                "top3_asin_conversion_share": maybe_number(cell(values, 22)),
                "suggested_bid_range": clean_text(cell(values, 24)),
            }
        )
    else:
        record.update(
            {
                "composite_score": maybe_number(cell(values, 6)),
                "monthly_search_volume": maybe_number(cell(values, 7)),
                "monthly_purchase_volume": maybe_number(cell(values, 8)),
                "purchase_rate": maybe_number(cell(values, 9)),
                "spr": maybe_number(cell(values, 10)),
                "ppc_bid": best_ppc_bid(values, 23, 11),
            }
        )

    if top_start:
        record.update(
            {
                "top1_asin": clean_text(cell(values, top_start)),
                "top1_click_share": maybe_number(cell(values, top_start + 1)),
                "top1_conversion_share": maybe_number(cell(values, top_start + 2)),
                "top2_asin": clean_text(cell(values, top_start + 3)),
                "top2_click_share": maybe_number(cell(values, top_start + 4)),
                "top2_conversion_share": maybe_number(cell(values, top_start + 5)),
                "top3_asin": clean_text(cell(values, top_start + 6)),
                "top3_click_share": maybe_number(cell(values, top_start + 7)),
                "top3_conversion_share": maybe_number(cell(values, top_start + 8)),
                "top10_asin": clean_text(cell(values, top_start + 9)),
            }
        )

    traffic_types = split_traffic_types(clean_text(record.get("traffic_word_type_raw")))
    record["keyword_tier"] = derive_keyword_tier(record)
    record["traffic_word_types"] = traffic_types
    record["raw_columns"] = raw_columns(raw_header_labels, values)
    record["verification_status"] = "local_source_only"
    record["risk_note"] = "Amazon/广告/搜索量字段来自本地模板样本,生产使用前需重新拉取最新平台数据"
    return record


def export_keywords(wb, sha: str) -> dict[str, Any]:
    ws = wb[KEYWORD_SHEET]
    raw_rows = used_rows(ws)
    if not raw_rows:
        raise RuntimeError(f"empty keyword sheet: {KEYWORD_SHEET}")
    _, raw_headers = raw_rows[0]
    _ = normalized_keyword_headers(raw_headers)
    raw_header_labels = [clean_text(v) for v in raw_headers]
    rows: list[dict[str, Any]] = []
    type_rows: list[dict[str, Any]] = []
    type_counter: Counter[str] = Counter()
    layout_counter: Counter[str] = Counter()
    tier_counter: Counter[str] = Counter()
    keyword_duplicates: Counter[str] = Counter()

    for ordinal, (source_row, values) in enumerate(raw_rows[1:], 1):
        record = build_keyword_record(ordinal, source_row, values, raw_header_labels, sha)
        keyword = clean_text(record.get("keyword"))
        if not keyword:
            continue
        rows.append(record)
        keyword_duplicates[keyword.lower()] += 1
        tier_counter[record["keyword_tier"]] += 1
        layout_counter[record["row_layout"]] += 1
        for traffic_type in record["traffic_word_types"]:
            type_counter[traffic_type] += 1
            type_rows.append(
                {
                    "keyword_asset_id": record["keyword_asset_id"],
                    "keyword": keyword,
                    "traffic_word_type": traffic_type,
                    "source_sheet": KEYWORD_SHEET,
                    "source_row": source_row,
                    "row_layout": record["row_layout"],
                }
            )

    full_headers = [
        "keyword_asset_id",
        "source_id",
        "source_file",
        "source_sha256_12",
        "source_sheet",
        "source_row",
        "row_layout",
        "keyword",
        "keyword_translation",
        "keyword_tier",
        "traffic_word_type_raw",
        "traffic_word_types",
        "trend_ref",
        "peak_season",
        "off_season",
        "composite_score",
        "monthly_search_volume",
        "monthly_purchase_volume",
        "purchase_rate",
        "spr",
        "ppc_bid",
        "traffic_share",
        "first_asin_organic_rank",
        "first_asin_ad_rank",
        "estimated_weekly_impressions",
        "related_products",
        "related_asin",
        "aba_weekly_rank",
        "title_density",
        "product_count",
        "supply_demand_ratio",
        "ad_competitor_count",
        "click_concentration",
        "top3_asin_conversion_share",
        "suggested_bid_range",
        "top1_asin",
        "top1_click_share",
        "top1_conversion_share",
        "top2_asin",
        "top2_click_share",
        "top2_conversion_share",
        "top3_asin",
        "top3_click_share",
        "top3_conversion_share",
        "top10_asin",
        "verification_status",
        "risk_note",
    ]
    csv_rows = []
    for row in rows:
        out = row.copy()
        out["traffic_word_types"] = "/".join(row.get("traffic_word_types", []))
        csv_rows.append(out)

    write_csv(OUT_DIR / "keyword_assets_full.csv", full_headers, csv_rows)
    write_jsonl(OUT_DIR / "keyword_assets_full.jsonl", rows)
    write_csv(
        OUT_DIR / "keyword_traffic_type_bridge.csv",
        ["keyword_asset_id", "keyword", "traffic_word_type", "source_sheet", "source_row", "row_layout"],
        type_rows,
    )
    type_summary_rows = [
        {"traffic_word_type": key, "keyword_count": count}
        for key, count in sorted(type_counter.items(), key=lambda x: (-x[1], x[0]))
    ]
    write_csv(OUT_DIR / "keyword_traffic_type_summary.csv", ["traffic_word_type", "keyword_count"], type_summary_rows)

    missing_core = defaultdict(int)
    for row in rows:
        for field in ("monthly_search_volume", "monthly_purchase_volume", "purchase_rate", "ppc_bid", "traffic_word_type_raw"):
            if row.get(field) in ("", None):
                missing_core[field] += 1

    duplicate_keywords = [key for key, count in keyword_duplicates.items() if count > 1]
    top_by_search = sorted(
        rows,
        key=lambda r: (to_number(r.get("monthly_search_volume")) or 0),
        reverse=True,
    )[:20]
    top_by_purchase = sorted(
        rows,
        key=lambda r: (to_number(r.get("monthly_purchase_volume")) or 0),
        reverse=True,
    )[:20]
    summary = {
        "keyword_sheet": KEYWORD_SHEET,
        "keyword_rows": len(rows),
        "source_sha256": sha,
        "source_sha256_12": sha[:12],
        "field_count": len(raw_header_labels),
        "row_layout_summary": dict(layout_counter),
        "traffic_type_summary": type_summary_rows,
        "keyword_tier_summary": dict(tier_counter),
        "missing_core_field_counts": dict(missing_core),
        "duplicate_keyword_count": len(duplicate_keywords),
        "top_by_monthly_search_volume": [
            {
                "keyword_asset_id": r["keyword_asset_id"],
                "keyword": r.get("keyword"),
                "monthly_search_volume": r.get("monthly_search_volume"),
                "monthly_purchase_volume": r.get("monthly_purchase_volume"),
                "traffic_word_type_raw": r.get("traffic_word_type_raw"),
                "row_layout": r.get("row_layout"),
            }
            for r in top_by_search
        ],
        "top_by_monthly_purchase_volume": [
            {
                "keyword_asset_id": r["keyword_asset_id"],
                "keyword": r.get("keyword"),
                "monthly_search_volume": r.get("monthly_search_volume"),
                "monthly_purchase_volume": r.get("monthly_purchase_volume"),
                "purchase_rate": r.get("purchase_rate"),
                "row_layout": r.get("row_layout"),
            }
            for r in top_by_purchase
        ],
    }
    write_json(OUT_DIR / "keyword_assets_summary.json", summary)
    return summary


def export_workbook() -> None:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(SOURCE_XLSX)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    sha = sha256(SOURCE_XLSX)
    wb_values = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)

    sheet_inventory = []
    for name in wb_values.sheetnames:
        ws = wb_values[name]
        rows = used_rows(ws)
        sheet_inventory.append(
            {
                "sheet_name": name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "used_row_count": len(rows),
                "used_column_count": len(rows[0][1]) if rows else 0,
                "is_keyword_asset_context": name == KEYWORD_SHEET or name in CONTEXT_SHEETS,
            }
        )

    keyword_summary = export_keywords(wb_values, sha)

    export_summaries = []
    for name in [KEYWORD_SHEET] + CONTEXT_SHEETS:
        if name in wb_values.sheetnames:
            export_summaries.append(export_sheet_csv(wb_values[name], OUT_DIR))

    write_csv(
        OUT_DIR / "sheet_inventory.csv",
        ["sheet_name", "max_row", "max_column", "used_row_count", "used_column_count", "is_keyword_asset_context"],
        sheet_inventory,
    )
    write_json(
        OUT_DIR / "workbook_profile.json",
        {
            "created_at": datetime.now(timezone.utc).isoformat(),
            "source_file": str(SOURCE_XLSX.relative_to(ROOT)),
            "source_id": SOURCE_ID,
            "source_sha256": sha,
            "source_sha256_12": sha[:12],
            "file_size_bytes": SOURCE_XLSX.stat().st_size,
            "sheet_count": len(wb_values.sheetnames),
            "sheet_inventory": sheet_inventory,
            "keyword_summary": keyword_summary,
            "export_summaries": export_summaries,
            "boundary": "local_source_only; no provider call; no Shopify store read/write",
        },
    )
    print(
        json.dumps(
            {
                "out": str(OUT_DIR),
                "source_sha256_12": sha[:12],
                "keyword_rows": keyword_summary["keyword_rows"],
                "field_count": keyword_summary["field_count"],
                "row_layout_summary": keyword_summary["row_layout_summary"],
                "sheet_exports": len(export_summaries),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    export_workbook()
